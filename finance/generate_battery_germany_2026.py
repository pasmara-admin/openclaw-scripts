import pandas as pd
from openpyxl import load_workbook

file_path = '/root/.openclaw/media/inbound/20260123_Overview_chemical_brands_Pasmara---7cfa565c-000f-4fe7-bdcb-f33c6acf832c.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# We need to find the row where System == 'Secondary Li-Ion < 50g'
# The data starts at row 2 in Excel.
# Col A (1): Brand
# Col B (2): Weight class
# Col C (3): Type group
# Col D (4): System
# Col E (5): Contract quantity in pieces
# Col F (6): Preis
# Col G (7): Weight per unit

for row in range(2, ws.max_row + 1):
    system_val = ws.cell(row=row, column=4).value
    if system_val and 'Secondary Li-Ion < 50g' in str(system_val):
        ws.cell(row=row, column=1, value='Bodyline Health and Massage, EcoXmas')
        ws.cell(row=row, column=5, value=15) # 12 + 3 forecast for the year
        ws.cell(row=row, column=7, value=40) # average weight in grams
        break

out_path = '/root/.openclaw/workspace-finance/Overview_chemical_brands_Pasmara_Filled.xlsx'
wb.save(out_path)
print("Saved to", out_path)
