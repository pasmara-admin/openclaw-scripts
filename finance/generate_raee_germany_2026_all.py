import pandas as pd
import pymysql
import re
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def get_connection():
    return pymysql.connect(
        host='34.38.166.212',
        user='john',
        password='3rmiCyf6d~MZDO41',
        database='kanguro'
    )

def identify_categories_de(df):
    for col in ['long_side', 'short_side', 'height']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    df['max_side'] = df[['long_side', 'short_side', 'height']].max(axis=1) * 100
    
    def classify(row):
        name = str(row['name']).lower()
        max_dim = row['max_side']
        weight_kg = pd.to_numeric(row['unit_weight_g'], errors='coerce') / 1000.0
        
        if max_dim == 0:
            if weight_kg > 10:
                max_dim = 100
            else:
                max_dim = 30
                
        # Exclusions
        if re.search(r'(pedana|parascintille|portalegna|supporto|palo|staffa|kit installazione|cornice|sassi|legnetti|pentola|padella|tegame|casseruola|batteria di pentole|copri|copertura|custodia|telo|ricambio|filtro)', name):
            return None
            
        if re.search(r'(sedia|sedie|sgabello|sgabelli|tavolo|tavoli|mobile|credenza|madia|poltrona|divano|letto|materasso|armadio|scrivania|mensola|coperchio|postazione|comodino)', name):
            if not re.search(r'(ventola|aspiratore|led|elettrica|massaggi)', name):
                return None

        # 1. Cooling and freezing / Temperature exchange
        if re.search(r'(climatizzatore|condizionatore|pompa di calore|deumidificatore|clima|fancoil)', name):
            return "Cat1"
            
        # 2. Screen & monitor devices
        if re.search(r'(televisore|monitor|schermo)', name) and not 'mobile' in name:
            return "Cat2"

        # 3. Lamps
        if re.search(r'(lampada|lampadario|applique|plafoniera|faretto|illuminazione|lanterna|luce|luci solari)', name) and not re.search(r'(specchio|mobile|ombrellone|valigetta|trolley|postazione|comodino)', name):
            return "Cat3"
            
        # 6. Small IT and telecommunication equipment
        if re.search(r'(power bank|batteria portatile|accumulatore)', name) and re.search(r'(litio|lithium|li-ion|li-po)', name):
            return "Cat6"

        # 4 & 5 Appliances
        is_appliance = re.search(r'(stufa|radiatore|caldaia|forno|lavatrice|asciugatrice|lavastoviglie|frigorifero|congelatore|piano cottura|cucina a gas|cucina elettrica|frullatore|friggitrice|mixer|tostapane|bollitore|macchina caffè|ferro da stiro|aspirapolvere|centrifuga|microonde|sandwich|waffle|griglia|frusta elettrica|stufetta|idromassaggio|asciugacapelli|rasoio|epilatore|tagliacapelli|piastra|spazzolino|massaggiatore|manicure|pedicure|fornetto|lampada uv|aspiratore unghie|doccia solare)', name)
        
        is_led_furniture = re.search(r'(poltrona|sedia|mobile|valigetta|postazione|comodino).*led|led.*(poltrona|sedia|mobile|valigetta|postazione|comodino)', name)
        is_electric_furniture = re.search(r'(poltrona|sedia|mobile|postazione|comodino).*(elettrica|massaggi|wireless usb|caricatore wireless)', name)
        
        if is_appliance or is_led_furniture or is_electric_furniture or re.search(r'(lampada|luce)', name):
            if max_dim > 50:
                return "Cat4"
            else:
                return "Cat5"
                
        return None

    df['Category'] = df.apply(classify, axis=1)
    return df

def main():
    conn = get_connection()
    
    # REMOVED: dp.is_active = b'1'
    query_products = """
    SELECT dp.id, dp.reference, dp.name, dp.long_side, dp.short_side, dp.height, dp.unit_weight_g, db.name as brand_name
    FROM dat_product dp
    LEFT JOIN dat_brand db ON dp.brand_id = db.id
    """
    df_products = pd.read_sql(query_products, conn)
    
    df_products = identify_categories_de(df_products)
    df_raee = df_products[df_products['Category'].notnull()].copy()
    
    # JAN 1st TO APR 30th 2026
    query_sales = """
    SELECT 
        sor.product_id,
        SUM(sor.qty) as total_qty
    FROM sal_order_row sor
    JOIN sal_order so ON sor.order_id = so.id
    WHERE so.date >= '2026-01-01' AND so.date <= '2026-04-30'
      AND so.delivery_country = 'Deutschland'
      AND so.source_srv = 'PS'
      AND so.state_id NOT IN ('09') 
      AND sor.is_deleted = b'0'
      AND so.is_deleted = b'0'
    GROUP BY sor.product_id
    """
    df_sales = pd.read_sql(query_sales, conn)
    
    query_refunds = """
    SELECT 
        bdr.product_id,
        SUM(bdr.qty) as refund_qty
    FROM bil_document_row bdr
    JOIN bil_document bd ON bdr.document_id = bd.id
    JOIN sal_order so ON bd.order_id = so.id
    WHERE bd.date >= '2026-01-01' AND bd.date <= '2026-04-30'
      AND so.delivery_country = 'Deutschland'
      AND so.source_srv = 'PS'
      AND bd.type_id IN (2, 4) 
      AND bd.is_deleted = 0
      AND bdr.is_deleted = 0
    GROUP BY bdr.product_id
    """
    df_refunds = pd.read_sql(query_refunds, conn)
    
    conn.close()
    
    df_final = df_raee.merge(df_sales, left_on='id', right_on='product_id', how='left')
    df_final = df_final.merge(df_refunds, left_on='id', right_on='product_id', how='left')
    
    df_final['total_qty'] = df_final['total_qty'].fillna(0)
    df_final['refund_qty'] = df_final['refund_qty'].fillna(0)
    
    df_final['Pezzi Venduti Netto'] = df_final['total_qty'] - df_final['refund_qty']
    df_final = df_final[df_final['Pezzi Venduti Netto'] > 0]
    
    if df_final.empty:
        print("No WEEE sales found for Germany Jan-Apr 2026 PS (including inactive).")
        return
        
    df_final['Peso Unitario (Kg)'] = df_final['unit_weight_g'].fillna(0) / 1000
    df_final['Peso Totale (kg)'] = df_final['Peso Unitario (Kg)'] * df_final['Pezzi Venduti Netto']
    
    df_final['brand_name'] = df_final['brand_name'].fillna('Produceshop')
    
    translations = {
        'Poltrona Massaggiante IRest Sl-A151 3D Massage Heaven': 'Massage Armchair IRest Sl-A151 3D Massage Heaven',
        'Poltrona Relax Elettrica con Sistema Alzapersona Per Anziani Giorgia Fx': 'Electric Relax Armchair with Lift System For Elderly Giorgia Fx',
        'Poltrona Relax Elettrica con Sistema Alzapersona e Ruote Per Anziani Giorgia': 'Electric Relax Armchair with Lift System and Wheels for Elderly Giorgia',
        'Ventola aspiratore polvere unghie per tavoli manicure Tornado': 'Nail dust aspirator fan for manicure tables Tornado',
        'Pannello colonna doccia in acciaio con miscelatore cascata idromassaggio Sirmione': 'Steel shower column panel with hydromassage waterfall mixer Sirmione',
        'Sedia gaming bianca poltrona LED reclinabile ergonomica cuscino Pixy': 'White ergonomic reclining gaming chair LED with pillow Pixy',
        'Sedia gaming bianca poltrona massaggiante LED reclinabile ergonomica Pixy Plus': 'White ergonomic reclining massage gaming chair LED Pixy Plus',
        'Poltrona relax elettrica reclinabile con alzapersona e ruote in Tessuto Marie per anziani': 'Electric reclining relax armchair with lift and wheels in Fabric Marie for elderly',
        'Lampada uv per unghie Led 48W portatile manicure pedicure Moon': 'Portable LED 48W UV nail lamp for manicure and pedicure Moon',
        'Tavolo manicure ricostruzione unghie aspiratore ruote cassettiera Sharpie': 'Manicure nail reconstruction table with aspirator wheels and drawer Sharpie',
        'Poltrona Relax Elettrica con Sistema Alzapersona Similpelle per Anziani Amalia Fix': 'Electric Relax Armchair with Lift System in Faux Leather for Elderly Amalia Fix',
        'Valigetta porta trucchi trolley make up con specchio LED casse audio Bluetooth Eva L': 'Makeup trolley case with LED mirror Bluetooth audio speakers Eva L',
        'Poltrona gaming ufficio LED RGB ergonomica reclinabile The Horde XL': 'Ergonomic office gaming chair LED RGB The Horde XL',
        'Poltrona gaming per bambini luci LED RGB sedia ergonomica Pixy Junior': 'Children gaming chair LED RGB Pixy Junior',
        'Poltrona sedia gaming ergonomica per bambini LED RGB The Horde junior': 'Children ergonomic gaming chair LED RGB The Horde junior',
        'Luci solari decorative esterno catena luminosa 200 LED giardino balcone Natale terrazzo NestX': 'Outdoor decorative solar lights 200 LED string garden balcony NestX',
        'Poltrona relax reclinabile elettrica con caricatore wireless USB Eleonora': 'Electric reclining relax armchair with USB wireless charger Eleonora',
        'Poltrona relax massaggiante e riscaldante ergonomica con poggiapiedi Myron': 'Ergonomic massaging and heating relax armchair with footrest Myron',
        'Lampada a sospensione moderna in cemento e vetro satinato Rock': 'Modern suspension lamp in concrete and frosted glass Rock',
        'Comodino moderno rotondo con ricarica wireless luce LED e 2 cassetti Moho': 'Modern round nightstand with wireless charging LED light and 2 drawers Moho',
        'Doccia solare da giardino esterno 25 litri con lavapiedi Arkema A120 Spring': 'Outdoor garden solar shower 25 liters with foot wash Arkema A120 Spring'
    }
    
    df_final['name_en'] = df_final['name'].map(lambda x: translations.get(x, x))
    
    file_path = '/root/.openclaw/media/inbound/20260123_Overview_WEEE_brands_and_categories_Pasmara---5e2e261f-54fb-43e7-aa80-1efa1166fd84.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active
    
    row_counters = {'Cat1': 3, 'Cat2': 3, 'Cat3': 3, 'Cat4': 3, 'Cat5': 3, 'Cat6': 3}
    
    df_final = df_final.sort_values(by=['Category', 'brand_name'])

    for index, row in df_final.iterrows():
        brand = row['brand_name']
        product = f"{row['reference']} - {row['name_en']} (Qty: {int(row['Pezzi Venduti Netto'])})"
        total_kg = row['Peso Totale (kg)']
        
        cat = row['Category']
        r = row_counters[cat]
        
        if cat == 'Cat1':
            c_brand, c_prod, c_kg = 1, 2, 3
        elif cat == 'Cat2':
            c_brand, c_prod, c_kg = 4, 5, 6
        elif cat == 'Cat3':
            c_brand, c_prod, c_kg = 7, 8, 9
        elif cat == 'Cat4':
            c_brand, c_prod, c_kg = 10, 11, 12
        elif cat == 'Cat5':
            c_brand, c_prod, c_kg = 13, 14, 15
        elif cat == 'Cat6':
            c_brand, c_prod, c_kg = 16, 17, 18
            
        ws.cell(row=r, column=c_brand, value=brand)
        ws.cell(row=r, column=c_prod, value=product)
        ws.cell(row=r, column=c_kg, value=total_kg)
        
        row_counters[cat] += 1

    out_path = '/root/.openclaw/workspace-finance/Overview_WEEE_brands_and_categories_Pasmara_Germany_Jan_Apr_2026_ALL.xlsx'
    wb.save(out_path)
    print("Saved to", out_path)
    print(df_final.groupby('Category')['Pezzi Venduti Netto'].sum())

if __name__ == "__main__":
    main()
